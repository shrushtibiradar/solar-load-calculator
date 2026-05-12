import easyocr
import re
from openpyxl import load_workbook
import os

reader = easyocr.Reader(['en'])

def extract_bill_data(file_path):

    results = reader.readtext(file_path, detail=0)

    text = " ".join(results)

    print("\n=========== OCR TEXT ===========\n")
    print(text)

    data = {
        "consumer_number": "",
        "units_consumed": "",
        "bill_amount": "",
        "load_kw": ""
    }
    consumer = re.search(r'(\d{12})', text)

    if consumer:
        data["consumer_number"] = consumer.group(1)
    amount = re.search(r'Rs\.?\s*(\d+\.?\d*)', text)

    if amount:
        data["bill_amount"] = amount.group(1)
    load = re.search(r'(\d+\.?\d*)\s*KW', text, re.I)

    if load:
        data["load_kw"] = load.group(1)
    if data["bill_amount"] != "":
        units = int(float(data["bill_amount"]) / 8)
        data["units_consumed"] = str(units)

    print("\n=========== EXTRACTED DATA ===========\n")
    print(data)

    return data


def fill_excel_template(template_path, output_path, data):

    print("\n=========== WRITING TO EXCEL ===========\n")

    wb = load_workbook(template_path)

    ws = wb.active

    ws["A1"] = "Consumer Number"
    ws["B1"] = data["consumer_number"]

    ws["A2"] = "Units Consumed"
    ws["B2"] = data["units_consumed"]

    ws["A3"] = "Bill Amount"
    ws["B3"] = data["bill_amount"]

    ws["A4"] = "Load KW"
    ws["B4"] = data["load_kw"]

    # Solar calculation
    if data["units_consumed"] != "":
        solar_size = int(data["units_consumed"]) / 120

        ws["A5"] = "Solar Size"
        ws["B5"] = round(solar_size, 2)

    print(ws["B1"].value)
    print(ws["B2"].value)
    print(ws["B3"].value)
    print(ws["B4"].value)

    wb.save(output_path)

    print("\nExcel Saved Successfully")